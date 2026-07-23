import streamlit as st
import urllib.request
import json
import re
import io
from pathlib import Path
from datetime import datetime, timedelta

import openpyxl

try:
    from streamlit_js_eval import streamlit_js_eval
    STOCKAGE_NAVIGATEUR_DISPO = True
except ImportError:
    STOCKAGE_NAVIGATEUR_DISPO = False

# --- 1. CONFIGURATION INITIALE DE LA PAGE ---
st.set_page_config(page_title="WeatherFly - Assistant Vol Libre", layout="wide")

# --- 1bis. PROFIL UTILISATEUR MÉMORISÉ DANS LE NAVIGATEUR (localStorage) ---
# Le spot et les compétences pilote sont sauvegardés dans le navigateur de chaque
# utilisateur et restaurés automatiquement à la connexion suivante.
CLES_PROFIL = ["vols_cumuls"] + [f"comp_{i}" for i in range(1, 17)] + ["sel_region", "sel_dept", "sel_spot"]


def charger_profil_navigateur():
    """Lit le profil stocké dans le navigateur et l'applique une seule fois par session."""
    if not STOCKAGE_NAVIGATEUR_DISPO or st.session_state.get("profil_applique"):
        return
    brut = streamlit_js_eval(
        js_expressions="localStorage.getItem('weatherfly_profil') || ''",
        key="chargement_profil",
    )
    if brut is None:
        return  # le navigateur n'a pas encore répondu, un rerun automatique suivra
    if brut:
        try:
            profil = json.loads(brut)
            for cle in CLES_PROFIL:
                if cle in profil and profil[cle] is not None:
                    st.session_state[cle] = profil[cle]
        except (json.JSONDecodeError, TypeError):
            pass
    st.session_state["profil_applique"] = True


def sauvegarder_profil_navigateur():
    """Écrit le profil courant dans le navigateur dès qu'une valeur change."""
    if not STOCKAGE_NAVIGATEUR_DISPO or not st.session_state.get("profil_applique"):
        return  # ne jamais écraser le profil stocké avant de l'avoir relu
    profil = {cle: st.session_state.get(cle) for cle in CLES_PROFIL}
    json_profil = json.dumps(profil, ensure_ascii=False)
    if st.session_state.get("_profil_sauvegarde") == json_profil:
        return
    streamlit_js_eval(
        js_expressions=f"localStorage.setItem('weatherfly_profil', {json.dumps(json_profil, ensure_ascii=False)})",
        key=f"sauvegarde_profil_{abs(hash(json_profil))}",
    )
    st.session_state["_profil_sauvegarde"] = json_profil

# --- 2. CHARGEMENT DE LA BASE DE DONNÉES SITES (EXCEL FFVL) ---
# Le fichier Excel doit être posé à côté de ce script.
# Colonnes attendues (feuille "Balises") : celles du fichier Balises_meteo_FFVL.xlsx
CHEMIN_EXCEL = Path(__file__).parent / "Balises_meteo_FFVL.xlsx"
NOM_FEUILLE = "Balises"

# Conversion des directions anglaises du fichier FFVL vers la notation française de l'appli
EN_VERS_FR = {"N": "N", "NE": "NE", "E": "E", "SE": "SE", "S": "S", "SW": "SO", "W": "O", "NW": "NO"}

# Le fichier FFVL utilise les anciennes régions (avant 2016) : conversion vers les régions actuelles
ANCIENNES_VERS_NOUVELLES_REGIONS = {
    "Alsace": "Grand Est",
    "Lorraine": "Grand Est",
    "Champagne-Ardenne": "Grand Est",
    "Aquitaine": "Nouvelle-Aquitaine",
    "Limousin": "Nouvelle-Aquitaine",
    "Poitou Charente": "Nouvelle-Aquitaine",
    "Auvergne": "Auvergne-Rhône-Alpes",
    "Rhône-Alpes": "Auvergne-Rhône-Alpes",
    "Basse Normandie": "Normandie",
    "Haute Normandie": "Normandie",
    "Bourgogne": "Bourgogne-Franche-Comté",
    "Franche-Comté": "Bourgogne-Franche-Comté",
    "Centre": "Centre-Val de Loire",
    "Ile de France": "Île-de-France",
    "Languedoc-Rousillon": "Occitanie",
    "Midi Pyrénées": "Occitanie",
    "Nord-Pas-De-Calais": "Hauts-de-France",
    "Picardie": "Hauts-de-France",
    "Paca": "Provence-Alpes-Côte d'Azur",
    "Pays de Loire": "Pays de la Loire",
    "Outre Mer": "Outre-Mer",
}
COMPASS_ANGLES = {"N": 0, "NE": 45, "E": 90, "SE": 135, "S": 180, "SO": 225, "O": 270, "NO": 315}
SANS_SPOT = "aucun"  # les lignes "aucun à <5 km" sont ignorées


def nettoyer_texte(txt):
    """Nettoie les textes bruts du fichier FFVL (retours chariot Excel, balises HTML)."""
    if not txt:
        return ""
    txt = str(txt).replace("_x000D_", " ")
    txt = re.sub(r"<[^>]*>", " ", txt)
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


ROSE_DES_VENTS = ["N", "NE", "E", "SE", "S", "SO", "O", "NO"]


def voisins_45_degres(direction):
    """Renvoie les deux directions à ±45° d'une direction donnée (ex : S -> SE, SO)."""
    idx = ROSE_DES_VENTS.index(direction)
    return [ROSE_DES_VENTS[(idx - 1) % 8], ROSE_DES_VENTS[(idx + 1) % 8]]


def parser_orientations(txt):
    """Transforme 'N·NW·SW' (notation FFVL) en ['N', 'NO', 'SO'] (notation appli).

    Accepte aussi la notation française (O, SO, NO) pour les colonnes remplies à la main.
    """
    if not txt:
        return []
    orientations = []
    for token in str(txt).replace(",", "·").replace(";", "·").split("·"):
        token = token.strip().upper()
        if token in EN_VERS_FR:
            token = EN_VERS_FR[token]
        if token in COMPASS_ANGLES and token not in orientations:
            orientations.append(token)
    return orientations


@st.cache_data(show_spinner="Chargement de la base de balises FFVL...")
def charger_base_spots(contenu_xlsx, _cle_cache):
    """Lit le fichier Excel FFVL et construit la hiérarchie Région > Département > Spot.

    Chaque spot est associé à sa meilleure balise : statut OK en priorité, puis la plus proche.
    """
    wb = openpyxl.load_workbook(io.BytesIO(contenu_xlsx), read_only=True, data_only=True)
    if NOM_FEUILLE not in wb.sheetnames:
        raise ValueError(f"Feuille '{NOM_FEUILLE}' introuvable dans le fichier Excel.")
    ws = wb[NOM_FEUILLE]
    lignes = ws.iter_rows(values_only=True)

    entete = [str(c).strip() if c is not None else "" for c in next(lignes)]
    col = {nom: i for i, nom in enumerate(entete)}

    colonnes_requises = ["Région", "Département", "Dépt", "N°", "Balise", "Statut",
                         "Lat", "Lon", "Spot le plus proche (<5 km)", "Dist. (km)",
                         "Vent idéal", "Vent possible"]
    manquantes = [c for c in colonnes_requises if c not in col]
    if manquantes:
        raise ValueError(f"Colonnes manquantes dans le fichier Excel : {', '.join(manquantes)}")

    def val(ligne, nom):
        i = col.get(nom)
        return ligne[i] if i is not None and i < len(ligne) else None

    hierarchie = {}
    for ligne in lignes:
        spot_brut = val(ligne, "Spot le plus proche (<5 km)")
        if not spot_brut or SANS_SPOT in str(spot_brut).lower():
            continue
        region = str(val(ligne, "Région") or "").strip()
        region = ANCIENNES_VERS_NOUVELLES_REGIONS.get(region, region)
        dept_nom = str(val(ligne, "Département") or "").strip()
        dept_code = str(val(ligne, "Dépt") or "").strip()
        lat, lon = val(ligne, "Lat"), val(ligne, "Lon")
        if not region or lat is None or lon is None:
            continue

        if dept_code.isdigit():
            dept_code = dept_code.zfill(2)
        dept_label = f"{dept_code} - {dept_nom}" if dept_code else dept_nom

        spot_nom = str(spot_brut).strip()
        spot_nom = spot_nom[0].upper() + spot_nom[1:]

        deco = parser_orientations(val(ligne, "Vent idéal"))
        # Vents favorables = colonne "Vent possible" + directions à ±45° des vents optimaux
        favorables_bruts = parser_orientations(val(ligne, "Vent possible"))
        for d_opt in deco:
            favorables_bruts.extend(voisins_45_degres(d_opt))
        deco_possible = []
        for d in favorables_bruts:
            if d not in deco and d not in deco_possible:
                deco_possible.append(d)
        orientations_connues = bool(deco or deco_possible)

        conseils = []
        for etiquette, colonne in [("Météo & pièges", "Météo & pièges"),
                                   ("Règles / restrictions", "Règles / restrictions"),
                                   ("Remarques", "Remarques")]:
            texte = nettoyer_texte(val(ligne, colonne))
            if texte:
                conseils.append(f"**{etiquette} :** {texte}")

        dist = val(ligne, "Dist. (km)")
        alt_deco = val(ligne, "Alt. déco (m)")
        alt_atterro = val(ligne, "Alt. atterro (m)")
        # Dénivelé déco -> atterro : calculé depuis les altitudes, sinon colonne D+ du fichier
        if alt_deco is not None and alt_atterro is not None:
            denivele = round(alt_deco - alt_atterro)
        else:
            d_plus = val(ligne, "D+ (m)")
            denivele = round(d_plus) if d_plus is not None else None

        candidat = {
            "lat": float(lat), "lon": float(lon),
            "deco": deco, "deco_possible": deco_possible,
            "orientations_connues": orientations_connues,
            "balise_ffvl_id": str(val(ligne, "N°") or "").strip(),
            "balise_nom": str(val(ligne, "Balise") or "").strip(),
            "balise_statut": str(val(ligne, "Statut") or "").strip(),
            "dist_km": float(dist) if dist is not None else None,
            "alt_deco": alt_deco,
            "alt_atterro": alt_atterro,
            "denivele": denivele,
            "thermique": str(val(ligne, "Therm.") or "").strip().lower() == "oui",
            "soaring": str(val(ligne, "Soar.") or "").strip().lower() == "oui",
            "conseil_site": "  \n".join(conseils) if conseils else "Pas d'information spécifique pour ce site dans la base FFVL.",
        }

        existant = hierarchie.setdefault(region, {}).setdefault(dept_label, {}).get(spot_nom)
        if existant is None or _est_meilleure_balise(candidat, existant):
            hierarchie[region][dept_label][spot_nom] = candidat

    # Tri alphabétique à tous les niveaux pour les menus déroulants
    return {
        region: {
            dept: dict(sorted(spots.items()))
            for dept, spots in sorted(depts.items())
        }
        for region, depts in sorted(hierarchie.items())
    }


def _est_meilleure_balise(candidat, existant):
    """Une balise OK bat une balise en maintenance ; à statut égal, la plus proche gagne."""
    def score(b):
        return (0 if b["balise_statut"] == "OK" else 1,
                b["dist_km"] if b["dist_km"] is not None else 9999)
    return score(candidat) < score(existant)


# --- 3. DICTIONNAIRES DE DÉCODAGE TEXTE DES CRITÈRES ---
def get_vols_comment(v):
    if v <= 15: return "(Vol guidé en école requis, début de l'apprentissage)"
    if v <= 35: return "(Phase d'autonomie initiale en conditions calmes)"
    if v <= 50: return "(Progression vers le brevet initial, ouverture aux premiers thermiques)"
    if v <= 100: return "(Pilote de site vu régulièrement, gestion des brises classiques)"
    if v <= 199: return "(Expérience solide, autonomie sur la majorité des sites connus)"
    return "(Volume de maturité : Expérience suffisante pour aborder tout type de site en sécurité)"

def get_skill_comment(val):
    if val <= 2: return "Niveau Initiation / Assistance indispensable"
    if val <= 5: return "Niveau Intermédiaire / Pratique en conditions calmes à modérées"
    if val <= 8: return "Niveau Avancé / Fluidité et bonnes parades en conditions classiques"
    return "Maîtrise absolue / Sérénité, rigueur et sécurité totale acquise"

# --- 4. FONCTIONS UTILITAIRES ---
def convertir_degres_en_direction(degres):
    if (degres >= 337.5) or (degres < 22.5): return "N"
    if 22.5 <= degres < 67.5: return "NE"
    if 67.5 <= degres < 112.5: return "E"
    if 112.5 <= degres < 157.5: return "SE"
    if 157.5 <= degres < 202.5: return "S"
    if 202.5 <= degres < 247.5: return "SO"
    if 247.5 <= degres < 292.5: return "O"
    return "NO"

def formater_fenetres(heures_valides, data_par_heure):
    if not heures_valides: return []
    precedent = heures_valides[0]
    blocs = []
    courant = [heures_valides[0]]
    for h in heures_valides[1:]:
        if h == precedent + 1:
            courant.append(h)
        else:
            blocs.append(courant)
            courant = [h]
        precedent = h
    blocs.append(courant)

    resultats_txt = []
    for bloc in blocs:
        h_debut = bloc[0]
        h_fin = bloc[-1]
        vitesses = [data_par_heure[h]["vitesse"] for h in bloc]
        indices = [data_par_heure[h]["indice"] for h in bloc]
        v_min, v_max = min(vitesses), max(vitesses)
        i_min, i_max = min(indices), max(indices)
        fourchette_v = f"{v_min} km/h" if v_min == v_max else f"{v_min}-{v_max} km/h"
        fourchette_i = f"{i_min}/10" if i_min == i_max else f"{i_min}-{i_max}/10"
        if h_debut == h_fin:
            resultats_txt.append(f"• {h_debut}:00 à {h_debut+1}:00 (Vent : {fourchette_v} | Agitation : {fourchette_i}) [Créneau court]")
        else:
            resultats_txt.append(f"• {h_debut}:00 à {h_fin+1}:00 (Vent : {fourchette_v} | Agitation : {fourchette_i})")
    return resultats_txt

def recuperer_vraie_meteo(lat, lon, date_str):
    url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&start_date={date_str}&end_date={date_str}&hourly=temperature_2m,wind_speed_10m,wind_gusts_10m,wind_direction_10m,precipitation,cape&wind_speed_unit=kmh&timezone=Europe%2FParis"
    try:
        req_om = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req_om, timeout=5) as response:
            data = json.loads(response.read().decode())
            return data.get("hourly", {})
    except Exception:
        return {}

def recuperer_donnees_balise_reelles(balise_id):
    url = f"https://www.balisemeteo.com/balise.php?idBalise={balise_id}"
    fallback = {
        "heure": "09:40", "vent_moyen": 5.0, "dir_moyen": "NC",
        "vent_max": 8.0, "dir_max": "NC", "indice": 1, "is_fallback": True
    }
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=4) as response:
            html_raw = response.read().decode('utf-8')
            text_clean = re.sub(r'<[^>]*>', ' ', html_raw)
            text_clean = re.sub(r'\s+', ' ', text_clean)

            match_heure = re.search(r"Relevé du \d{2}/\d{2}/\d{4} - (\d{2}:\d{2})", text_clean)
            heure_reelle = match_heure.group(1) if match_heure else "09:40"

            dir_moyen = "NC"
            vent_moyen = 5.0
            match_moyen = re.search(r"Vent moyen Direction\s*:\s*(.+?)\s*Vitesse\s*:\s*([\d\.,]+)", text_clean, re.IGNORECASE)
            if match_moyen:
                dir_moyen = match_moyen.group(1).strip()
                vent_moyen = float(match_moyen.group(2).replace(',', '.'))

            dir_max = "NC"
            vent_max = vent_moyen + 3.0
            match_maxi = re.search(r"Vent maxi Direction\s*:\s*(.+?)\s*Vitesse\s*:\s*([\d\.,]+)", text_clean, re.IGNORECASE)
            if match_maxi:
                dir_max = match_maxi.group(1).strip()
                vent_max = float(match_maxi.group(2).replace(',', '.'))

            delta_rafale = max(0.0, vent_max - vent_moyen)
            indice = min(10, round((delta_rafale / 3) + (vent_moyen / 5)))

            return {
                "heure": heure_reelle,
                "vent_moyen": vent_moyen, "dir_moyen": dir_moyen,
                "vent_max": vent_max, "dir_max": dir_max,
                "indice": max(1, indice), "is_fallback": False
            }
    except Exception:
        return fallback

# --- 5. INTERFACE GRAPHIQUE STREAMLIT ---
st.title("WeatherFly - Assistant Vol Libre")

if "refresh_counter" not in st.session_state:
    st.session_state.refresh_counter = 0

aujourd_hui_str = datetime.now().strftime("%Y-%m-%d")

# Chargement de la base Excel : fichier local prioritaire, sinon upload manuel
contenu_xlsx = None
cle_cache = None
if CHEMIN_EXCEL.exists():
    contenu_xlsx = CHEMIN_EXCEL.read_bytes()
    cle_cache = f"{CHEMIN_EXCEL}|{CHEMIN_EXCEL.stat().st_mtime}"
else:
    st.warning(f"Fichier '{CHEMIN_EXCEL.name}' introuvable à côté du script. Charge-le manuellement :")
    fichier_upload = st.file_uploader("Base de balises FFVL (.xlsx)", type=["xlsx"])
    if fichier_upload is not None:
        contenu_xlsx = fichier_upload.getvalue()
        cle_cache = f"upload|{fichier_upload.name}|{len(contenu_xlsx)}"

if contenu_xlsx is None:
    st.stop()

try:
    SPOTS_HIERARCHIE = charger_base_spots(contenu_xlsx, cle_cache)
except Exception as e:
    st.error(f"Impossible de lire la base de balises : {e}")
    st.stop()

nb_spots = sum(len(spots) for depts in SPOTS_HIERARCHIE.values() for spots in depts.values())
st.caption(f"📚 Base FFVL chargée : {nb_spots} spots répartis sur {len(SPOTS_HIERARCHIE)} régions.")

# Restauration du profil mémorisé (spot + compétences) depuis le navigateur
charger_profil_navigateur()
st.session_state.setdefault("vols_cumuls", 10)

col_gauche, col_droite = st.columns([3, 2])

with col_gauche:
    # --- PROFIL PILOTE ---
    st.subheader("👤 1. Calculateur de Niveau Pilote")

    # Curseur adaptatif : pas de 1 unité de 0 à 50, puis pas de 5 unités jusqu'à 200+
    pas_vols = 1 if st.session_state["vols_cumuls"] <= 50 else 5
    vols = st.slider("Volume de vols cumulés :", min_value=0, max_value=200, step=pas_vols, key="vols_cumuls")

    st.caption(get_vols_comment(vols))

    notes_competences = []

    with st.expander("🪁 Techniques au Sol"):
        c1 = st.slider("Gonflage dos voile :", 0, 10, key="comp_1", help="De 0 (guidage requis) à 10 (pente raide/espace restreint sans vent)")
        st.caption(get_skill_comment(c1))
        c2 = st.slider("Gonflage face voile :", 0, 10, key="comp_2", help="De 0 (non pratiqué) à 10 (jeu au sol instinctif dans les rafales)")
        st.caption(get_skill_comment(c2))
        c3 = st.slider("Gonflage technique & vent fort :", 0, 10, key="comp_3", help="De 0 (danger potentiel) à 10 (haute montagne/déco engagé ou technique)")
        st.caption(get_skill_comment(c3))
        c4 = st.slider("Pré-vol et check-list de sécurité :", 0, 10, key="comp_4", help="De 0 (oublis fréquents) à 10 (rigueur absolue contre les automatismes)")
        st.caption(get_skill_comment(c4))
        notes_competences.extend([c1, c2, c3, c4])

    with st.expander("🪂 Techniques en Vol"):
        c5 = st.slider("Gestion du plan d'approche (PTU / PTR) :", 0, 10, key="comp_5", help="De 0 (guidage requis) à 10 (précision chirurgicale tout gradient)")
        st.caption(get_skill_comment(c5))
        c6 = st.slider("Atterrissage hors terrain (Vachage) :", 0, 10, key="comp_6", help="De 0 (à proscrire) à 10 (adaptation immédiate sur zone inconnue)")
        st.caption(get_skill_comment(c6))
        c7 = st.slider("Virages coordonnés (roulis/tangage) :", 0, 10, key="comp_7", help="De 0 (sur-pilotage fréquent) à 10 (pilotage instinctif fluide)")
        st.caption(get_skill_comment(c7))
        c8 = st.slider("Pilotage actif en turbulences :", 0, 10, key="comp_8", help="De 0 (crispation aux commandes) à 10 (vol fluide en air agité)")
        st.caption(get_skill_comment(c8))
        c9 = st.slider("Gestion des incidents (fermetures) :", 0, 10, key="comp_9", help="De 0 (panique/pas de réflexe) à 10 (calme olympien, bagage SIV complet)")
        st.caption(get_skill_comment(c9))
        c10 = st.slider("Exploitation du soaring :", 0, 10, key="comp_10", help="De 0 (vol balistique) à 10 (optimisation du relief par vent faible ou fort)")
        st.caption(get_skill_comment(c10))
        c11 = st.slider("Exploitation des thermiques :", 0, 10, key="comp_11", help="De 0 (simple traversée) à 10 (analyse complète et montée sereine)")
        st.caption(get_skill_comment(c11))
        c12 = st.slider("Pilotage aux arrières & accélérateur :", 0, 10, key="comp_12", help="De 0 (source d'inquiétude) à 10 (utilisation reflexe de toute la plage)")
        st.caption(get_skill_comment(c12))
        notes_competences.extend([c5, c6, c7, c8, c9, c10, c11, c12])

    with st.expander("🧠 Analyse & Sécurité"):
        c13 = st.slider("Analyse de la manche à air :", 0, 10, key="comp_13", help="De 0 (lecture ardue) à 10 (décodage complet de l'environnement)")
        st.caption(get_skill_comment(c13))
        c14 = st.slider("Analyse météo et aérologique :", 0, 10, key="comp_14", help="De 0 (dépendance aux icônes) à 10 (compréhension fine émagramme/pièges)")
        st.caption(get_skill_comment(c14))
        c15 = st.slider("Planification de vol (Cheminement) :", 0, 10, key="comp_15", help="De 0 (vol local strict) à 10 (gestion des espaces et secours permanents)")
        st.caption(get_skill_comment(c15))
        c16 = st.slider("Prise de décision & Renoncement :", 0, 10, key="comp_16", help="De 0 (effet mouton) à 10 (sagesse aéronautique, savoir dire non acquis)")
        st.caption(get_skill_comment(c16))
        notes_competences.extend([c13, c14, c15, c16])

    # Calcul algorithmique ajusté : vols à 100% dès 200 vols (score max = 3.0 pts)
    avg_skills = sum(notes_competences) / len(notes_competences)
    vols_score = min(3.0, (vols / 200) * 3)

    final_score = (avg_skills * 0.7) + vols_score
    if vols <= 10 and avg_skills == 0:
        final_score = 0.0
    final_score = min(10.0, final_score)
    note_pilote = max(1, round(final_score))

    # Application du nouveau classement centré sur l'autonomie de sécurité
    if final_score < 0.5:
        rank = "Grand Débutant"
    elif final_score < 3.0:
        rank = "Autonomie Initiale (Niveau 1-2)"
    elif final_score < 5.0:
        rank = "Autonomie Confirmée / Brevet Initial (Niveau 3-4)"
    elif final_score < 7.0:
        rank = "Pilote de Site / Brevet Pilote (Niveau 5-6)"
    else:
        rank = "Autonome sur tout type de site en toute condition météo sans se mettre en danger"

    st.info(f"📊 **Statut :** {rank}  \n🎯 **Note finale calculée :** {final_score:.1f}/10 (Arrondie à : {note_pilote}/10)")

    st.markdown("---")

    # --- SPOT & DATE ---
    st.subheader("🧭 2. Localisation du Spot & Date")
    # Si une valeur mémorisée n'existe plus (ex : base Excel mise à jour), on la
    # remplace par le premier choix valide : une affectation explicite garde le
    # widget synchronisé, contrairement à une suppression de la clé.
    regions = list(SPOTS_HIERARCHIE.keys())
    if st.session_state.get("sel_region") not in regions:
        st.session_state["sel_region"] = regions[0]
    region_selectionnee = st.selectbox("Région :", regions, key="sel_region")

    departements = list(SPOTS_HIERARCHIE[region_selectionnee].keys())
    if st.session_state.get("sel_dept") not in departements:
        st.session_state["sel_dept"] = departements[0]
    dept_selectionne = st.selectbox("Département :", departements, key="sel_dept")

    spots = list(SPOTS_HIERARCHIE[region_selectionnee][dept_selectionne].keys())
    if st.session_state.get("sel_spot") not in spots:
        st.session_state["sel_spot"] = spots[0]
    spot_name = st.selectbox("Site officiel :", spots, key="sel_spot")

    spot_config = SPOTS_HIERARCHIE[region_selectionnee][dept_selectionne][spot_name]

    dates_possibles = [(datetime.now() + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(15)]
    date_selectionnee = st.selectbox("Date du vol :", dates_possibles)

    st.markdown("---")
    analyser_clic = st.button("RECHERCHER ET ANALYSER", type="primary")

    # --- VERDICT METEO ---
    st.subheader("Verdict Météo & Aérologie")
    st.warning(
        "⚠️ **Avertissement :** cette analyse météo automatique ne dispense pas de faire son plan de vol, "
        "de s'informer sur les spécificités du site et de vérifier la météo réelle une fois sur place. "
        "Dans le doute, demandez des informations aux pilotes locaux — et si le doute persiste, "
        "remettez votre vol à une autre fois."
    )
    indice_preve_actuel = 1

    if analyser_clic:
        with st.spinner("Interrogation des serveurs météo..."):
            hourly_data = recuperer_vraie_meteo(spot_config["lat"], spot_config["lon"], date_selectionnee)

        if hourly_data and "time" in hourly_data:
            heures_valides_int = []
            data_par_heure = {}
            facteurs_limitants = set()
            historique_vents = []

            vent_max_autorise = 15 if note_pilote <= 3 else (20 if note_pilote <= 6 else 26)
            seuil_agitation_max = 6 if note_pilote <= 3 else (8 if note_pilote <= 6 else 10)

            heure_courante = datetime.now().hour
            axes_acceptes = spot_config["deco"] + spot_config["deco_possible"]

            for i in range(len(hourly_data["time"])):
                heure_texte = hourly_data["time"][i].split("T")[1][:5]
                heure_int = int(heure_texte.split(":")[0])
                if heure_int < 8 or heure_int > 21: continue

                vitesse = round(hourly_data["wind_speed_10m"][i])
                v_rafales = round(hourly_data["wind_gusts_10m"][i]) if "wind_gusts_10m" in hourly_data else vitesse
                direction = convertir_degres_en_direction(hourly_data["wind_direction_10m"][i])
                pluie = hourly_data["precipitation"][i]
                cape_val = hourly_data.get("cape", [0])[i] if "cape" in hourly_data else 0

                delta_rafale = max(0, v_rafales - vitesse)
                indice_agitation = min(10, round((delta_rafale / 3) + (vitesse / 5) + (cape_val / 200)))

                if heure_int == heure_courante:
                    indice_preve_actuel = max(1, indice_agitation)

                heure_bloquee = False
                cause_heure = ""
                meteo_heure = f"{direction}, {vitesse} km/h"

                if indice_agitation > seuil_agitation_max:
                    cause_heure = f"☀️ Agitation thermique ({indice_agitation}/10)"
                    facteurs_limitants.add(f"☀️ Aérologie trop agitée ({indice_agitation}/10 > {seuil_agitation_max}/10)")
                    heure_bloquee = True
                elif c9 < 5 and indice_agitation >= 8:
                    cause_heure = f"🔥 Thermique marqué (Maîtrise incidents requise)"
                    facteurs_limitants.add("🔥 Pic thermique (Maîtrise des fermetures/incidents insuffisante)")
                    heure_bloquee = True

                if heure_bloquee:
                    historique_vents.append(f"• {heure_texte} ({meteo_heure}) : {cause_heure}")
                    continue

                if pluie > 0.1:
                    cause_heure = f"🌧️ Pluie ({pluie} mm)"
                    facteurs_limitants.add("🌧️ Risque de précipitations")
                    heure_bloquee = True
                elif vitesse > vent_max_autorise:
                    cause_heure = f"💨 Trop fort ({vitesse} km/h)"
                    facteurs_limitants.add(f"💨 Vitesse du vent supérieure à ton maximum autorisé ({vent_max_autorise} km/h)")
                    heure_bloquee = True
                elif vitesse > 15 and c2 < 5:
                    cause_heure = f"🛑 Face voile requis ({vitesse} km/h)"
                    facteurs_limitants.add("🛑 Gonflage face voile insuffisant (Requis dès 15 km/h)")
                    heure_bloquee = True
                elif vitesse > 5 and spot_config["orientations_connues"] and direction not in axes_acceptes:
                    cause_heure = f"🧭 Vent travers/cul ({direction})"
                    facteurs_limitants.add(f"🧭 Vent travers/cul ({direction}) : hors des vents optimaux et favorables du site")
                    heure_bloquee = True

                if heure_bloquee:
                    historique_vents.append(f"• {heure_texte} ({meteo_heure}) : {cause_heure}")
                else:
                    heures_valides_int.append(heure_int)
                    data_par_heure[heure_int] = {"vitesse": vitesse, "indice": indice_agitation}

            liste_fenetres = formater_fenetres(heures_valides_int, data_par_heure)
            if liste_fenetres:
                st.success("🟢 FEU VERT POUR LE VOL")
                for f in liste_fenetres: st.write(f)
                if not spot_config["orientations_connues"]:
                    st.warning("🧭 Orientations de décollage inconnues dans la base : l'axe du vent n'a PAS été contrôlé. Vérifie l'orientation du déco sur place.")
            else:
                st.error("🛑 FEU ROUGE : RESTE AU SOL")
                for cause in facteurs_limitants: st.write(f"• {cause}")

            if historique_vents:
                st.markdown("**Détail des heures bloquées :**")
                for h_hist in historique_vents: st.write(h_hist)
        else:
            st.warning("Impossible de récupérer les prévisions Open-Meteo.")

with col_droite:
    # --- BLOC VERROUILLÉ : RÈGLES ---
    st.subheader("Règles de Sécurité Intégrales")
    st.markdown("""
    *   **Niveau Débutant (Note 1-3) :** Max **15 km/h** | Agitation max **6/10**. Le thermique fort est à proscrire.
    *   **Niveau Progression (Note 4-6) :** Max **20 km/h** | Agitation max **8/10**. Maîtrise des petites turbulences.
    *   **Niveau Confirmé (Note 7-10) :** Max **26 km/h** | Agitation max **10/10**. Gestion des conditions fortes et autonomie toutes masses d'air.
    *   **Règle du Face Voile :** Niveau intermédiaire requis sur le curseur Sol (Note >= 5) si le vent moyen dépasse **15 km/h** pour assurer un décollage en sécurité.
    *   **Règle des Incidents :** Interdiction de voler si l'indice d'agitation prévu atteint **8/10** sans une note minimale de 5 en gestion des incidents (SIV/Fermetures).
    """)

    # --- BLOC : SPÉCIFICITÉS DU SPOT (issues du fichier Excel FFVL) ---
    st.markdown("---")
    st.subheader("📌 Spécificités du Spot")

    if spot_config["deco"]:
        st.write(f"• **Vents optimaux :** {', '.join(spot_config['deco'])}")
    if spot_config["deco_possible"]:
        st.write(f"• **Vents favorables :** {', '.join(spot_config['deco_possible'])}")
    if not spot_config["orientations_connues"]:
        st.write("• **Vents :** orientations non renseignées dans la base ⚠️")

    infos_alt = []
    if spot_config["alt_deco"]: infos_alt.append(f"Déco {round(spot_config['alt_deco'])} m")
    if spot_config["alt_atterro"]: infos_alt.append(f"Atterro {round(spot_config['alt_atterro'])} m")
    if spot_config["denivele"]: infos_alt.append(f"Dénivelé {spot_config['denivele']} m")
    if infos_alt:
        st.write(f"• **Altitudes :** {' | '.join(infos_alt)}")

    pratiques = []
    if spot_config["thermique"]: pratiques.append("Thermique ☀️")
    if spot_config["soaring"]: pratiques.append("Soaring 🌬️")
    if pratiques:
        st.write(f"• **Pratiques :** {', '.join(pratiques)}")

    st.markdown(spot_config["conseil_site"])

    # --- BLOC : RELEVÉ RÉEL BALISE ---
    ffvl_id = spot_config.get("balise_ffvl_id")
    if ffvl_id:
        st.markdown("---")
        st.subheader("📡 Relevé Réel BaliseMétéo")
        detail_balise = f"Balise **{spot_config['balise_nom']}** (n°{ffvl_id})"
        if spot_config["dist_km"] is not None:
            detail_balise += f" — à {spot_config['dist_km']} km du spot"
        st.write(detail_balise)
        if spot_config["balise_statut"] != "OK":
            st.warning(f"⚠️ Balise signalée « {spot_config['balise_statut']} » dans la base : le relevé en direct peut être indisponible ou faux.")
        st.markdown(f"[Accéder à la page de la balise FFVL n°{ffvl_id}](https://www.balisemeteo.com/balise.php?idBalise={ffvl_id})")

        if date_selectionnee == aujourd_hui_str:
            if st.button("🔄 Rafraîchir la balise", key="refresh_balise"):
                st.session_state.refresh_counter += 1

            balise_reelle = recuperer_donnees_balise_reelles(ffvl_id)

            if balise_reelle.get("is_fallback"):
                st.error("❌ Relevé en direct indisponible (les valeurs ci-dessous sont des valeurs par défaut, ne pas s'y fier).")

            st.write(f"• **Heure du relevé en direct :** {balise_reelle['heure']}")
            st.write(f"• **Vent moyen constaté :** {balise_reelle['vent_moyen']} km/h ({balise_reelle['dir_moyen']})")
            st.write(f"• **Vent max constaté :** {balise_reelle['vent_max']} km/h ({balise_reelle['dir_max']})")
            st.write(f"• **Indice d'agitation réel :** {balise_reelle['indice']}/10")

            if 'hourly_data' in locals() and hourly_data:
                diff_indice = balise_reelle['indice'] - indice_preve_actuel
                if diff_indice > 0:
                    st.warning(f"📈 **Comparaison :** Les conditions réelles sur site sont plus fortes que prévu (**+{diff_indice} point(s)** d'indice d'agitation par rapport à la météo).")
                elif diff_indice < 0:
                    st.info(f"📉 **Comparaison :** Les conditions réelles sur site sont plus calmes que prévu (**{diff_indice} point(s)** d'indice d'agitation par rapport à la météo).")
                else:
                    st.success("🎯 **Comparaison :** L'agitation réelle sur site est parfaitement conforme aux prévisions météo (0 point d'écart).")
        else:
            st.info("ℹ️ Les données en temps réel et le bouton de rafraîchissement ne sont disponibles que pour le jour J.")

# --- 6. SAUVEGARDE DU PROFIL DANS LE NAVIGATEUR ---
sauvegarder_profil_navigateur()
